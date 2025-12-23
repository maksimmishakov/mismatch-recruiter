"""Excel export utilities for analytics and candidates data."""

import io
from datetime import datetime

def export_analytics_to_excel(analytics_data):
    """Export analytics data to Excel bytes.
    
    Args:
        analytics_data: Dictionary with analytics metrics
        
    Returns:
        Excel file as bytes
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return None
    
    wb = Workbook()
    ws = wb.active
    ws.title = 'Analytics'
    
    # Header styling
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws['A1'] = 'Analytics Dashboard Report'
    ws['A1'].font = Font(bold=True, size=14, color='4472C4')
    ws.merge_cells('A1:D1')
    ws['A2'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
    ws['A2'].font = Font(italic=True, size=10)
    
    # Summary section
    row = 4
    headers = ['Metric', 'Value', '', '']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row, col, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    row = 5
    metrics = [
        ('Total Resumes', analytics_data.get('total_candidates', 0)),
        ('Average Score', f"{analytics_data.get('average_score', 0):.1f}%"),
        ('Status: Approved', analytics_data.get('status_breakdown', {}).get('approved', 0)),
        ('Status: Rejected', analytics_data.get('status_breakdown', {}).get('rejected', 0)),
        ('Status: Pending', analytics_data.get('status_breakdown', {}).get('pending', 0)),
    ]
    
    for metric, value in metrics:
        ws.cell(row, 1, metric).border = border
        ws.cell(row, 2, value).border = border
        row += 1
    
    # Skills section
    row += 2
    ws.cell(row, 1, 'Top Skills').font = Font(bold=True, size=11, color='4472C4')
    row += 1
    
    skill_headers = ['Skill', 'Count']
    for col, header in enumerate(skill_headers, 1):
        cell = ws.cell(row, col, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    
    row += 1
    top_skills = analytics_data.get('top_skills', [])
    for skill, count in top_skills[:10]:
        ws.cell(row, 1, skill).border = border
        ws.cell(row, 2, count).border = border
        row += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

def export_candidates_to_excel(candidates):
    """Export candidates data to Excel bytes.
    
    Args:
        candidates: List of candidate dictionaries
        
    Returns:
        Excel file as bytes
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return None
    
    wb = Workbook()
    ws = wb.active
    ws.title = 'Candidates'
    
    # Header styling
    header_fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = ['ID', 'Name', 'Email', 'Position', 'Score', 'Status', 'Skills', 'Date Added']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Data rows
    for row, candidate in enumerate(candidates, 2):
        ws.cell(row, 1, candidate.get('id')).border = border
        ws.cell(row, 2, candidate.get('name')).border = border
        ws.cell(row, 3, candidate.get('email')).border = border
        ws.cell(row, 4, candidate.get('position')).border = border
        ws.cell(row, 5, candidate.get('score')).border = border
        ws.cell(row, 6, candidate.get('status')).border = border
        ws.cell(row, 7, ', '.join(candidate.get('skills', []))).border = border
        ws.cell(row, 8, candidate.get('date_added')).border = border
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 30
    ws.column_dimensions['H'].width = 18
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

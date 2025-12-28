"""Report Generator Service - Export analytics data to various formats (CSV, Excel, PDF)."""
from datetime import datetime
from typing import Dict, Optional
import csv
import json
from io import StringIO
import os
import logging

logger = logging.getLogger(__name__)


class ReportGenerator:
    """Generate and export analytics reports in multiple formats."""
    
    def __init__(self, export_dir: str = "exports/"):
        self.export_dir = export_dir
        if not os.path.exists(export_dir):
            os.makedirs(export_dir)
    
    def generate_csv(self, analytics_data: Dict) -> str:
        """Convert analytics data to CSV format"""
        try:
            output = StringIO()
            writer = csv.writer(output)
            writer.writerow(['Metric', 'Value', 'Generated Date'])
            for key, value in analytics_data.items():
                writer.writerow([key, value, datetime.now().isoformat()])
            
            csv_content = output.getvalue()
            filename = f"report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
            filepath = os.path.join(self.export_dir, filename)
            
            with open(filepath, 'w') as f:
                f.write(csv_content)
            
            logger.info(f"CSV report generated: {filename}")
            return filename
        except Exception as e:
            logger.error(f"Error generating CSV: {str(e)}")
            raise
    
    def generate_excel(self, analytics_data: Dict) -> str:
        """Convert analytics data to Excel format"""
        try:
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font
            except ImportError:
                logger.error("openpyxl not installed")
                raise ImportError("openpyxl package required")
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Analytics"
            ws['A1'] = 'Metric'
            ws['B1'] = 'Value'
            ws['C1'] = 'Generated Date'
            
            for cell in ['A1', 'B1', 'C1']:
                ws[cell].font = Font(bold=True)
            
            row = 2
            for key, value in analytics_data.items():
                ws[f'A{row}'] = str(key)
                ws[f'B{row}'] = value
                ws[f'C{row}'] = datetime.now().isoformat()
                row += 1
            
            ws.column_dimensions['A'].width = 30
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 25
            
            filename = f"report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath = os.path.join(self.export_dir, filename)
            wb.save(filepath)
            
            logger.info(f"Excel report generated: {filename}")
            return filename
        except Exception as e:
            logger.error(f"Error generating Excel: {str(e)}")
            raise
    
    def generate_pdf(self, analytics_data: Dict, title: str = "Analytics Report") -> str:
        """Convert analytics data to PDF format"""
        try:
            try:
                from reportlab.lib.pagesizes import letter
                from reportlab.pdfgen import canvas
            except ImportError:
                logger.error("reportlab not installed")
                raise ImportError("reportlab package required")
            
            filename = f"report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            filepath = os.path.join(self.export_dir, filename)
            c = canvas.Canvas(filepath, pagesize=letter)
            width, height = letter
            
            c.setFont("Helvetica-Bold", 16)
            c.drawString(50, height - 50, title)
            c.setFont("Helvetica", 10)
            c.drawString(50, height - 70, f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            
            c.setFont("Helvetica", 11)
            y = height - 110
            c.setFont("Helvetica-Bold", 11)
            c.drawString(50, y, "Metric")
            c.drawString(300, y, "Value")
            y -= 20
            c.setFont("Helvetica", 10)
            
            for key, value in analytics_data.items():
                if y < 50:
                    c.showPage()
                    y = height - 50
                    c.setFont("Helvetica", 10)
                
                c.drawString(50, y, str(key)[:40])
                c.drawString(300, y, str(value)[:30])
                y -= 15
            
            c.setFont("Helvetica", 8)
            c.drawString(50, 30, f"Page 1 of 1 | {datetime.now().isoformat()}")
            c.save()
            logger.info(f"PDF report generated: {filename}")
            return filename
        except Exception as e:
            logger.error(f"Error generating PDF: {str(e)}")
            raise
    
    def generate_json(self, analytics_data: Dict) -> str:
        """Export analytics data as JSON"""
        try:
            report = {
                "generated_at": datetime.now().isoformat(),
                "metrics": analytics_data,
                "total_metrics": len(analytics_data)
            }
            
            filename = f"report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
            filepath = os.path.join(self.export_dir, filename)
            
            with open(filepath, 'w') as f:
                json.dump(report, f, indent=2)
            
            logger.info(f"JSON report generated: {filename}")
            return filename
        except Exception as e:
            logger.error(f"Error generating JSON: {str(e)}")
            raise
    
    def cleanup_old_reports(self, days_to_keep: int = 30) -> str:
        """Delete reports older than specified days"""
        try:
            from datetime import timedelta
            import glob
            
            cutoff_date = datetime.now() - timedelta(days=days_to_keep)
            deleted_count = 0
            
            for pattern in ["report_*.csv", "report_*.xlsx", "report_*.pdf", "report_*.json"]:
                for filepath in glob.glob(os.path.join(self.export_dir, pattern)):
                    file_mtime = datetime.fromtimestamp(os.path.getmtime(filepath))
                    if file_mtime < cutoff_date:
                        os.remove(filepath)
                        deleted_count += 1
            
            message = f"Cleaned up {deleted_count} reports older than {days_to_keep} days"
            logger.info(message)
            return message
        except Exception as e:
            logger.error(f"Error in cleanup: {str(e)}")
            raise
    
    def get_report_size(self, filename: str) -> Optional[float]:
        """Get report file size in kilobytes"""
        try:
            filepath = os.path.join(self.export_dir, filename)
            if os.path.exists(filepath):
                size_bytes = os.path.getsize(filepath)
                size_kb = round(size_bytes / 1024, 2)
                return size_kb
            return None
        except Exception as e:
            logger.error(f"Error getting file size: {str(e)}")
            return None
